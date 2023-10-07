# SEO环境： schedule
import os
import requests
import pandas as pd
import numpy as np
import jsonpath
import json
import time
import datetime
import xlwings as xw
import telebot
import hmac, base64, struct, hashlib
from openpyxl import Workbook, load_workbook
from openpyxl import formatting, styles
from openpyxl.styles import Color, PatternFill, Font, Border
from PIL import ImageGrab
import pyperclip
import warnings
warnings.filterwarnings('ignore')

day = 0
pages_user = 150
pages_fircharge = 60
access_token = '121.3b699a76ba3f0e0a1a920e929e0be12a.Y__S7vw4TPZXbg-CvtT5SKdFTIijP7cBKPrRQdw.F4UX0w'
# access_token ='121.1e832791a57b87542b2bb51e2f3f5bfa.Y_Uhf0W55kh6mBiTGZX0qWg0O5ZqJYZmPyHTqi8.HEyD3w'
start_date = (datetime.datetime.now()+datetime.timedelta(days=day)).strftime('%Y%m%d')
end_date = (datetime.datetime.now()+datetime.timedelta(days=day)).strftime('%Y%m%d')

url = 'http://fundmng.bsportsadmin.com/api/manage/fund/withdraw/record/list/history'
session = requests.session()

# 第一次获取token
submit_url = 'http://fundmng.bsportsadmin.com/api/manage/user/admin/login/submit'
header0 = {
    'Accept':'application/json, text/plain, */*',
    # 'Accept-Encoding':'gzip, deflate',
    'Accept-Language':'zh-CN,zh;q=0.9',
    'Connection':'keep-alive',
    'Content-Length':'48',
    'Content-Type':'application/x-www-form-urlencoded',
    'Cookie':'admin-uid=690; admin-token=db76bebda5274c80adaadd40bd794f24',
    'Device_id':'1.0',
    'Gl_version':'2.0',
    'Host':'fundmng.bsportsadmin.com',
    'Language':'zh_CN',
    'Origin':'http://fundmng.bsportsadmin.com',
    'Os_type':'0',
    'Referer':'http://fundmng.bsportsadmin.com/login',
    'Sign':'2bc4c378817f47731f0adf450a627d19',
    'Some':'header',
    'Systemid':"",
    'Timestamp':'1692415901000',
    'Token':'-1',
    'Uid':'-1',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36',
    'Version':'1.0'
}
def get_google_code(secret):
    key = base64.b32decode(secret, True)
    msg = struct.pack(">Q", int(time.time()) // 30)
    google_code = hmac.new(key, msg, hashlib.sha1).digest()
    # 很多网上的代码不可用，就在于这儿，没有chr字符串
    o = ord(chr(google_code[19])) & 15
    # google_code = (struct.unpack(">I", google_code[o:o + 4])[0] & 0x7fffffff) % 1000000
    google_code = (struct.unpack(">I", google_code[o:o + 4])[0] & 0x7fffffff) % 1000000
    return '%06d' % google_code

# 获取今日数据
print('启动百度统计API----')
shuju_website = {'domain':[],
                 '日期':[],
                 'pv':[],
                 'uv':[],
                 'ip':[]}
qishi = {'domain':[],
         '日期':[],
         '时间':[],
         'pv':[],
         'uv':[],
         'ip':[]}

url_siteid = f'https://openapi.baidu.com/rest/2.0/tongji/config/getSiteList?access_token={access_token}'
response = requests.get(url_siteid)

dic_website = {}
for k,v in zip(jsonpath.jsonpath(json.loads(response.text),'$..domain'),jsonpath.jsonpath(json.loads(response.text),'$..site_id')):
    dic_website[k]=v
# with open(r'C:\Users\User\Desktop\SEO\12-18\dic_website.txt','r') as f:
#     dic_website=f.read()
# 分别获取各网站数据
app = xw.App(visible=False,add_book=False)
book = app.books.open(r'C:\Users\User\Desktop\SEO\12-18\今日数据(py接口).xlsx')
sheet1 = book.sheets['网站概况']
# sheet1.range('A2:E2').expand('down').clear_contents()
sheet_qishu = book.sheets['趋势分析']
# sheet_qishu.range('A2:F2').expand('down').clear_contents()

session_web = requests.Session()

for k in dic_website:
    url_web = f'https://openapi.baidu.com/rest/2.0/tongji/report/getData?access_token={access_token}&site_id={dic_website[k]}&method=overview/getTimeTrendRpt&start_date={start_date}&end_date={end_date}&metrics=pv_count,visitor_count,ip_count'
    response = session_web.get(url_web)
    response.encoding='utf8'
    # 趋势数据
    for i in range(24):
        qishi['domain'].append(k)
        qishi['日期'].append((datetime.datetime.now()+datetime.timedelta(days=day)).strftime('%Y/%m/%d'))
        qishi['时间'].append(i)
        qishi['pv'].append(json.loads(response.text)['result']['items'][1][i][0])
        qishi['uv'].append(json.loads(response.text)['result']['items'][1][i][1])
        qishi['ip'].append(json.loads(response.text)['result']['items'][1][i][2])
    result_pv_uv_ip = []
    # 遍历列表并相加元素
    for i in range(3):
        sum = 0
        for j in range(len(json.loads(response.text)['result']['items'][1])):
            try:
                sum += json.loads(response.text)['result']['items'][1][j][i]
            except:
                sum +=0
        result_pv_uv_ip.append(sum)

    shuju_website['domain'].append(k)
    shuju_website['日期'].append((datetime.datetime.now()+datetime.timedelta(days=day)).strftime('%Y/%m/%d'))
    shuju_website['pv'].append(result_pv_uv_ip[0])
    shuju_website['uv'].append(result_pv_uv_ip[1])
    shuju_website['ip'].append(result_pv_uv_ip[2])
    time.sleep(1)
sheet1.range('A2').options(index=False,header = False).value = pd.DataFrame(shuju_website)
sheet_qishu.range('A2').options(index=False,header = False).value = pd.DataFrame(qishi)
book.save()
app.quit()
print('今日数据获取完毕！')

# 后续采集会员列表，首充记录、数据处理
# 读取今日数据，及历史数据
data_today = pd.read_excel(r'C:\Users\User\Desktop\SEO\12-18\今日数据(py接口).xlsx')
data_2_today = pd.read_excel(r'C:\Users\User\Desktop\SEO\12-18\今日数据(py接口).xlsx','趋势分析')
daili = pd.read_excel(r'C:\Users\User\Desktop\SEO\数据+ip历史.xlsx','代理总表')
his_data  = pd.read_excel(r'C:\Users\User\Desktop\SEO\SEO总表(12点+18点).xlsx','数据(18点)_3')

# 采集会员列表和会员存记录
url_fircharge = 'http://fundmng.bsportsadmin.com/api/manage/data/detail/firstRecharge'
url_user = 'http://fundmng.bsportsadmin.com/api/manage/user/maintain/user/list'

# 采集首存报表
dic_fir = dict({'会员名':[], '所属代理':[],'注册时间':[], '交易时间':[], '交易类型':[], '币种':[], '金额':[]})
# 昨天开始时间戳
yesterday = datetime.date.today() + datetime.timedelta(days=day)
yesterday_start_time = int(time.mktime(time.strptime(str(yesterday), '%Y-%m-%d')))
# 昨天结束时间戳
yesterday_end_time = int(time.mktime(time.strptime(str(datetime.date.today()+ datetime.timedelta(days=day+1)), '%Y-%m-%d'))) - 1
print(yesterday_start_time)
print(yesterday_end_time)

#---------------------token----------------------
# 采集token
google_code = get_google_code('64ehnxj6yily5bhv23kgb62ozuh6yuu2')
data0 = {
    'username': 'Marquis',
    'password': 'qwer123456',
    'code': google_code
}
session0 = requests.Session()
response0  =session0.post(url=submit_url,data=data0,headers=header0)
response0.encoding = 'utf-8'
obj0 = json.loads(response0.text)
token = obj0['data']['token']

header = {
    'Accept':'application/json, text/plain, */*',
    # 'Accept-Encoding':'gzip, deflate',
    'Accept-Language':'zh-CN,zh;q=0.9',
    'Connection':'keep-alive',
    'Content-Length':'75',
    'Content-Type':'application/x-www-form-urlencoded',
    'Cookie':'admin-token=67c8b1bd1b434f898ed8570a860355b8; admin-uid=690',
    'Device_id':'1.0',
    'Gl_version':'2.0',
    'Host':'fundmng.bsportsadmin.com',
    'Language':'zh_CN',
    'Menuid':'100112',
    'Opeartionmenu':'%u62A5%u8868%u67E5%u8BE2-%u4F1A%u5458%u9996%u5B58%u62A5%u8868',
    'Origin':'http://fundmng.bsportsadmin.com',
    'Os_type':'0',
    'Referer':'http://fundmng.bsportsadmin.com/system/report-query/report-first-recharge',
    'Sign':'ca83944852acc68fe114cbc65f1e1d22',
    'Some':'header',
    'Systemid':'54',
    'Timestamp':'1692092554000',
    'Token':token,
    'Uid':'690',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36',
    'Version':'1.0'
}
for page in range(1,pages_fircharge+1):
    data = {
        'page': page,
        'size': 20,
        'tradeType': 0,
        'isFake': 0,
        'stime': yesterday_start_time*1000,
        'etime': yesterday_end_time*1000+999
    }
    response = session.post(url_fircharge,headers=header,data=data,timeout=300)
    response.encoding = 'utf-8'
    obj = json.loads(response.text)

    for i in obj['data']['list']:
        dic_fir['会员名'].append(i['userName'])
        dic_fir['所属代理'].append(i['parentName'])
        dic_fir['注册时间'].append(time.strftime('%Y/%m/%d %H:%M:%S',time.localtime(i['createTime']//1000)))
        dic_fir['交易时间'].append(time.strftime('%Y/%m/%d %H:%M:%S',time.localtime(i['regTime']//1000)))
        dic_fir['交易类型'].append(i['tradeType'])
        dic_fir['币种'].append(i['coinCode'])
        dic_fir['金额'].append(i['amount'])
firChargeUser = pd.DataFrame(dic_fir)
print('会员首存行列：',firChargeUser.shape)

# 采集会员列表
dic_user = dict({'会员账号':[], '姓名':[],'代理':[], '注册时间':[], '备注':[]})
for page in range(1,pages_user+1):
    data2 = {
        'page':page,
        'size':20,
        'userVip':'0,1,2,3,4,5,6,7,8,9,10,11',
        'status':'0,1,2,4',
        'sortType':'3',
        'sortStr':'descend',
        'searchType':'1',
        'channelId':'34',
        'registeredStartDate':yesterday_start_time*1000,
        'registeredEndDate':yesterday_end_time*1000+999,
    }
    response2 = session.post(url_user,headers=header,data=data2,timeout=300)
    response2.encoding = 'utf-8'
    obj2 = json.loads(response2.text)
    for i in obj2['data']['list']:
        dic_user['会员账号'].append(i['username'])
        if i['reallyName'] !='':
            dic_user['姓名'].append(i['reallyName'])
        else:
            dic_user['姓名'].append('--')

        dic_user['代理'].append(i['parentName'])
        dic_user['注册时间'].append(time.strftime('%Y/%m/%d %H:%M:%S',time.localtime(i['registerDate']//1000)))
        dic_user['备注'].append(i['remark'])
user = pd.DataFrame(dic_user)
print('用户列表行列:',user.shape)
# 删除测试账号
user = user[~user['会员账号'].str.contains('test')&~user['会员账号'].str.contains('ceshi')&~user['姓名'].str.contains('测试') \
            &~user['姓名'].str.contains('cheshi')&~user['代理'].str.contains('测试')&~user['代理'].str.contains('cheshi') \
            &~user['备注'].str.contains('测试')&~user['备注'].str.contains('试玩')&~user['备注'].str.contains('晒单')]
print('去重后：',user.shape)

print('开始处理shuju...')
shuju = pd.DataFrame({'人员':['Paddy', 'Tony', 'Max', 'Martin', 'Zed', 'Hugo', 'Aber', 'DK', 'Ben','当日汇总'],
                      '日期':(datetime.datetime.now()+datetime.timedelta(days=day)).strftime('%Y/%m/%d'),
                      '时间':(datetime.datetime.now()+datetime.timedelta(days=day)).strftime('%H:%M'),
                      '发送IP':0,
                      '接收IP':0,
                      '对比昨天(总IP)':0,
                      '对比前3天均值(总IP)':0,
                      '对比前7天均值(总IP)':0,
                      '对比昨天(接收IP)':0,
                      '对比前3天均值(接收IP)':0,
                      '对比前7天均值(接收IP)':0,
                      '对比昨天(总注册)':0,
                      '对比前3天均值(总注册)':0,
                      '对比前7天均值(总注册)':0,
                      '对比昨天(总开户)':0,
                      '对比前3天均值(总开户)':0,
                      '对比前7天均值(总开户)':0})

shuju.set_index('人员',inplace = True)
data_today.set_index('网站名(domain)',inplace=True)
# data_today['IP']=pd.to_numeric(data_today['IP'],errors='coerce').replace(np.nan,0).astype('int64')
# data_today=data_today.groupby('网站名(domain)').agg({'IP':sum})
try:
    shuju.loc['Paddy','发送IP']=data_today.loc['paddy.com','IP']
except:
    shuju.loc['Paddy','发送IP']=0
try:
    shuju.loc['Paddy','接收IP']=data_today.loc['paddy.bty','IP']
except:
    shuju.loc['Paddy','接收IP']=0
try:
    shuju.loc['Tony','发送IP']=data_today.loc['tonyb.com','IP']
except:
    shuju.loc['Tony', '发送IP'] = 0
try:
    shuju.loc['Tony','接收IP']=data_today.loc['tony.bty','IP']
except:
    shuju.loc['Tony', '接收IP'] = 0
try:
    shuju.loc['Max','发送IP']=data_today.loc['mulu.com','IP']
except:
    shuju.loc['Max', '发送IP'] = 0
try:
    shuju.loc['Max','接收IP']=data_today.loc['max.bty','IP']
except:
    shuju.loc['Max', '接收IP'] = 0
try:
    shuju.loc['Martin','发送IP']=data_today.loc['redquan.com','IP']
except:
    shuju.loc['Martin', '发送IP'] = 0
try:
    shuju.loc['Martin','接收IP']=data_today.loc['martin.bty','IP']
except:
    shuju.loc['Martin','接收IP']=0
try:
    shuju.loc['Zed','发送IP']=data_today.loc['zed.com','IP']
except:
    shuju.loc['Zed', '发送IP'] = 0
try:
    shuju.loc['Zed','接收IP']=data_today.loc['zed.bty','IP']
except:
    shuju.loc['Zed', '接收IP'] = 0
try:
    shuju.loc['Hugo','发送IP']=data_today.loc['hugo.com','IP']
except:
    shuju.loc['Hugo', '发送IP'] = 0
try:
    shuju.loc['Hugo','接收IP']=data_today.loc['hugo.bty','IP']
except:
    shuju.loc['Hugo','接收IP']=0
try:
    shuju.loc['Aber','发送IP']=data_today.loc['aber.com','IP']
except:
    shuju.loc['Aber','发送IP']=0
try:
    shuju.loc['Aber','接收IP']=data_today.loc['aber.bty','IP']
except:
    shuju.loc['Aber','接收IP']=0
try:
    shuju.loc['DK','发送IP']=data_today.loc['dk.com','IP']
except:
    shuju.loc['DK','发送IP']=0
try:
    shuju.loc['DK','接收IP']=data_today.loc['dk.bty','IP']
except:
    shuju.loc['DK','接收IP']=0
try:
    shuju.loc['Ben','发送IP']=data_today.loc['ben.com','IP']
except:
    shuju.loc['Ben','发送IP']=0
try:
    shuju.loc['Ben','接收IP']=data_today.loc['ben.bty','IP']
except:
    shuju.loc['Ben','接收IP']=0
shuju.loc['当日汇总','发送IP']=shuju['发送IP'].sum()
shuju.loc['当日汇总','接收IP']=shuju['接收IP'].sum()

shuju['日期'] = pd.to_datetime(shuju['日期'])
shuju.insert(1,'人员2',shuju.index)
shuju['人员2']=shuju['人员2'].str.lower()

# 第1次merge前，重置索引
shuju.reset_index(inplace=True)

merge_user = pd.merge(user,daili,how = 'left',left_on='代理',right_on='代理线')
data_todaySEO = merge_user.groupby('seo变化数据团队').agg({'seo变化数据团队':len})
data_todaySEO.rename(columns={'seo变化数据团队':'注册'},inplace=True)
data_todaySEO.reset_index(inplace=True)
data_todaySEO['人员2'] = data_todaySEO['seo变化数据团队'].str.lower()
data_todaySEO.set_index('seo变化数据团队',inplace=True)

shuju=shuju.merge(data_todaySEO,on='人员2',how='left')

shuju['注册率(%)'] = round(shuju['注册']/shuju['接收IP']*100,2)

merge_charge = pd.merge(firChargeUser,daili,how='left',left_on='所属代理',right_on='代理线')
data_todayCHARGE = merge_charge.groupby('seo变化数据团队').agg({'seo变化数据团队':len})
data_todayCHARGE= data_todayCHARGE.rename(columns={'seo变化数据团队':'开户'})
data_todayCHARGE.reset_index(inplace=True)
data_todayCHARGE['seo变化数据团队']=data_todayCHARGE['seo变化数据团队'].str.lower()
data_todayCHARGE= data_todayCHARGE.rename(columns={'seo变化数据团队':'人员2'})
# 第2次merge
shuju = pd.merge(shuju,data_todayCHARGE,how='left',on='人员2')
shuju['转化率(%)'] = round(shuju['开户']/shuju['注册']*100,2)


#------------
# his_data  = pd.read_csv(r'C:\Users\User\Desktop\SEO\SEO每日更新_814.csv',encoding='gbk')
his_data['日期']= pd.to_datetime(his_data['日期'])
be_data = his_data[his_data['日期']==(shuju['日期'][0]+datetime.timedelta(days=-1))][:-1]

shuju.fillna(0,inplace=True)
shuju.set_index('人员',inplace = True)
shuju.sort_index(inplace=True)
be_data.set_index('人员',inplace=True)
be_data.sort_index(inplace=True)

be3_data = his_data[his_data['日期']>=(shuju['日期'][0]+datetime.timedelta(days=-3))]
be3_data = be3_data.groupby('人员').mean()[:-1]

be7_data = his_data[his_data['日期']>=(shuju['日期'][0]+datetime.timedelta(days=-7))]
be7_data = be7_data.groupby('人员').mean()[:-1]

shuju['对比昨天(总IP)']=shuju['发送IP']-be_data['总IP']
shuju['对比前3天均值(总IP)']= shuju['发送IP']-be3_data['总IP']
shuju['对比前7天均值(总IP)']= shuju['发送IP']-be7_data['总IP']

shuju['对比昨天(接收IP)']=shuju['接收IP']-be_data['接收IP']
shuju['对比前3天均值(接收IP)']= shuju['接收IP']-be3_data['接收IP']
shuju['对比前7天均值(接收IP)']= shuju['接收IP']-be7_data['接收IP']

shuju['对比昨天(总注册)']=shuju['注册']-be_data['注册']
shuju['对比前3天均值(总注册)']= shuju['注册']-be3_data['注册']
shuju['对比前7天均值(总注册)']= shuju['注册']-be7_data['注册']

shuju['对比昨天(总开户)']=shuju['开户']-be_data['开户']
shuju['对比前3天均值(总开户)']= shuju['开户']-be3_data['开户']
shuju['对比前7天均值(总开户)']= shuju['开户']-be7_data['开户']

shuju = shuju.iloc[:,:5].join(shuju.iloc[:,-4:]).join(shuju.iloc[:,5:-4])
shuju.fillna(0,inplace=True)
for name in shuju.index:
    if shuju.loc[name,'注册']==0:
        shuju.loc[name,'转化率(%)']=shuju.loc[name,'开户']*100
    if shuju.loc[name,'接收IP']==0:
        shuju.loc[name,'注册率(%)']=shuju.loc[name,'注册']*100

shuju.loc[:,'对比昨天(总IP)':]=shuju.loc[:,'对比昨天(总IP)':].astype('int64')
shuju['注册'] = shuju['注册'].astype('int64')
shuju['开户'] = shuju['开户'].astype('int64')

for i in shuju.iloc[:,5:].columns:
    shuju.loc['当日汇总',i]= shuju[i].sum()

# 重置三个率
shuju.loc['当日汇总','注册率(%)']=round(shuju.loc['当日汇总','注册']/shuju.loc['当日汇总','接收IP']*100,2)
shuju.loc['当日汇总','转化率(%)']=round(shuju.loc['当日汇总','开户']/shuju.loc['当日汇总','注册']*100,2)

shuju.insert(3,'人员',shuju.index)
shuju.drop('人员2',inplace=True,axis=1)
print('shuju处理完成。。。。')

# 重置历史数据
be_data = his_data[his_data['日期']==(shuju['日期'][0]+datetime.timedelta(days=-1))]
shuju2 = shuju.copy()
shuju2= shuju2.rename(columns={'开户':'开户2','注册':'注册2','接收IP':'接收IP2','对比昨天(总开户)':'开户','对比昨天(总注册)':'注册','对比昨天(接收IP)':'接收IP','对比昨天(总IP)':'总IP'})

# 写入发送txt文本
aip_7 = list(shuju[:-1].loc[shuju['对比前7天均值(总IP)']<-99,:].index)
rip_7 = list(shuju[:-1].loc[shuju['对比前7天均值(接收IP)']<-99,:].index)
zhuce_7 = list(shuju[:-1].loc[shuju['对比前7天均值(总注册)']<-4,:].index)
acc_7 = list(shuju[:-1].loc[shuju['对比前7天均值(总开户)']<-4,:].index)
zhuanhuali = list(shuju[:-1].loc[shuju[:-1]["转化率(%)"]<30,:]["人员"])
d_zhuce = int(shuju.loc["当日汇总","注册"]-be_data.loc[be_data["人员"]=="当日汇总","注册"].values[0])
d_acc = int(shuju.loc["当日汇总","开户"]-be_data.loc[be_data["人员"]=="当日汇总","开户"].values[0])
d_zhuanhua = round(shuju.loc["当日汇总","转化率(%)"]-be_data.loc[be_data["人员"]=="当日汇总","转化率(%)"].values[0],2)

with open(r'C:\Users\User\Desktop\SEO\截图文件\seo_18-3.txt','w') as f:
    f.write('#SEO激活监控18点\n')
    f.write(f'截止今日18点,   注册:  {shuju.loc["当日汇总","注册"]} ,开户:  {shuju.loc["当日汇总","开户"]}，整体'
            f'转化率 : {shuju.loc["当日汇总","转化率(%)"]}%\n')
    f.write(f"对比昨日18点,   注册:  {int(be_data.loc[be_data['人员']=='当日汇总','注册'].values[0])} ,开户:  {int(be_data.loc[be_data['人员']=='当日汇总','开户'].values[0])}，整体转化率 : {be_data.loc[be_data['人员']=='当日汇总','转化率(%)'].values[0]}%\n")
    f.write(f'同比昨日,  ')
    if d_zhuce>0:
        f.write(f'注册 上升：{abs(d_zhuce)} 个,')
    if d_zhuce<0:
        f.write(f'注册 下降：{abs(d_zhuce)} 个,')
    if d_zhuce==0:
        f.write(f'注册 无变化')
    if d_acc>0:
        f.write(f'开户 上升：{abs(d_acc)} 个,')
    if d_acc<0:
        f.write(f'开户 下降：{abs(d_acc)} 个,')
    if d_acc==0:
        f.write(f'开户 无变化')
    if d_zhuanhua>0:
        f.write(f'转化率 上升：{abs(d_zhuanhua)}%')
    if d_zhuanhua<0:
        f.write(f'转化率 下降：{abs(d_zhuanhua)}%')
    if d_zhuanhua==0:
        f.write(f'转化率 无变化')
    f.write('\n')
    f.write('\n')
    f.write('人员注册下降：\n')
    for i in list(shuju.loc[shuju['对比昨天(总注册)']<0,]['人员'])[:-1]:
        f.write(f'{i} : {abs(shuju.loc[i,"对比昨天(总注册)"])}\n')
    f.write('\n')
    f.write('人员开户下降：\n')
    for i in list(shuju.loc[shuju['对比昨天(总开户)']<0,]['人员'])[:-1]:
        f.write(f'{i} : {abs(shuju.loc[i,"对比昨天(总开户)"])}\n')
    f.write('\n')
    f.write(f'对比前7天均值明显下降:\n')
    if len(aip_7)>0:
        f.write(f'总IP：')
        for i in aip_7:
            f.write(f'{i}{", "}')
        f.write('\n')
    if len(rip_7)>0:
        f.write(f'接收IP：')
        for i in rip_7:
            f.write(f'{i}{", "}')
        f.write('\n')
    if len(zhuce_7)>0:
        f.write(f'总注册：')
        for i in zhuce_7:
            f.write(f'{i}{", "}')
        f.write('\n')
    if len(acc_7)>0:
        f.write(f'总开户：')
        for i in acc_7:
            f.write(f'{i}{", "}')
        f.write('\n')
    f.write('\n')
    if len(zhuanhuali)>0:
        f.write(f'转化率<30%：')
        for i in zhuanhuali:
            f.write(f'{i}{", "}')

# 增加行末表头

# 增加%
# shuju['注册率(%)'] =shuju['注册率(%)'].apply(lambda x: str(x)+'%')
# shuju['转化率(%)'] =shuju['转化率(%)'].apply(lambda x: str(x)+'%')

# shuju = shuju.append(header_shuju)
# 保存数据
app = xw.App(visible=False,add_book=False)
book = app.books.open(r'C:\Users\User\Desktop\SEO\SEO总表(12点+18点).xlsx')

sheet_shuju = book.sheets['数据(18点)_3']
row_shuju = sheet_shuju.used_range.last_cell.row

sheet_shuju['A'+str(row_shuju+1)].options(index=False,header = False).value = shuju
# sheet_ip['A'+str(row_ip+1)].options(index=False,header = False).value = ip_data
sheet_tem = book.sheets['临时']
shuju3 = shuju.copy()
shuju3['注册率(%)'] =shuju3['注册率(%)'].apply(lambda x: str(x)+'%')
shuju3['转化率(%)'] =shuju3['转化率(%)'].apply(lambda x: str(x)+'%')
sheet_tem['A3'].options(index=False,header = False).value = shuju3
book.save()
book.close()

# # 添加条件格式
wb = load_workbook(r'C:\Users\User\Desktop\SEO\SEO总表(12点+18点).xlsx')
ws = wb['数据(18点)_3']
ws_tem = wb['临时']
redFill = Font(color='FF0000')
ws.conditional_formatting.add(f'J{row_shuju +1}:U{row_shuju+10}',
                              formatting.rule.CellIsRule(operator='lessThan',
                                                         formula=['0'],
                                                         font=redFill))
# 临时表的条件格式
ws_tem.conditional_formatting.add(f'J3:U12',
                                  formatting.rule.CellIsRule(operator='lessThan',
                                                             formula=['0'],
                                                             font=redFill))
wb.save(filename=r'C:\Users\User\Desktop\SEO\SEO总表(12点+18点).xlsx')
wb.close()
# # 保存截图
book2 = app.books.open(r'C:\Users\User\Desktop\SEO\SEO总表(12点+18点).xlsx')
tem_shuju = book2.sheets['临时']
# sheet2_shuju = book2.sheets['数据(12点)_3']
#
# 截图
pyperclip.copy('')
range_shuju = tem_shuju.range('A1:U12')
range_shuju.api.CopyPicture()
img_shuju = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
img_shuju.save(r'C:\Users\User\Desktop\SEO\截图文件\shuju(18h)-3.png')  # 保存图片
# pyperclip.copy('')

time.sleep(2)
#
book2.save()
book2.close()
app.quit()
# # 发送到群
with open(r'C:\Users\User\Desktop\SEO\截图文件\seo_18-3.txt','r') as f:
    text = f.read()

# bot_DA = telebot.TeleBot("6106076754:AAHjxPSBpyjwpY-lq1iEslUufW46XQvAfr0")
# bot_DA.send_photo(-677235937,open(r'C:\Users\User\Desktop\SEO\截图文件\shuju(18h)-3.png','rb'))
# # bot_DA.send_message(-677235937,'#SEO激活监控12点')
# bot_DA.send_message(-677235937,text)
# bot_DA.stop_polling()
# 查看  -677235937, -812533282  鲲鹏流量： -321785338

# print('发送完毕。')

