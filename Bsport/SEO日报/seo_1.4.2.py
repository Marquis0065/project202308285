import os
import warnings
warnings.filterwarnings('ignore')
import requests
import pandas as pd
import numpy as np
import jsonpath
import json
import time
import datetime
import xlwings as xw
import telebot
import hmac, base64, struct, hashlib
from openpyxl import Workbook, load_workbook
from openpyxl import formatting, styles
from openpyxl.styles import Color, PatternFill, Font, Border
from PIL import ImageGrab
import pyperclip

pd.set_option('display.max_colwidth', None) #显示单元格完整信息
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)

day = -1
start_date = (datetime.datetime.now()+datetime.timedelta(days=day)).strftime('%Y%m%d')
end_date = (datetime.datetime.now()+datetime.timedelta(days=day)).strftime('%Y%m%d')
last_date = (datetime.datetime.now()+datetime.timedelta(days=day-1)).strftime('%Y%m%d')
pages_user = 150
pages_fircharge = 60
access_token = '121.1e832791a57b87542b2bb51e2f3f5bfa.Y_Uhf0W55kh6mBiTGZX0qWg0O5ZqJYZmPyHTqi8.HEyD3w'
# access_token = '121.3b699a76ba3f0e0a1a920e929e0be12a.Y__S7vw4TPZXbg-CvtT5SKdFTIijP7cBKPrRQdw.F4UX0w'

url = 'http://fundmng.bsportsadmin.com/api/manage/fund/withdraw/record/list/history'
session = requests.session()

# 第一次获取token
submit_url = 'http://fundmng.bsportsadmin.com/api/manage/user/admin/login/submit'
header0 = {
    'Accept':'application/json, text/plain, */*',
    # 'Accept-Encoding':'gzip, deflate',
    'Accept-Language':'zh-CN,zh;q=0.9',
    'Connection':'keep-alive',
    'Content-Length':'48',
    'Content-Type':'application/x-www-form-urlencoded',
    'Cookie':'admin-uid=690; admin-token=db76bebda5274c80adaadd40bd794f24',
    'Device_id':'1.0',
    'Gl_version':'2.0',
    'Host':'fundmng.bsportsadmin.com',
    'Language':'zh_CN',
    'Origin':'http://fundmng.bsportsadmin.com',
    'Os_type':'0',
    'Referer':'http://fundmng.bsportsadmin.com/login',
    'Sign':'2bc4c378817f47731f0adf450a627d19',
    'Some':'header',
    'Systemid':"",
    'Timestamp':'1692415901000',
    'Token':'-1',
    'Uid':'-1',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36',
    'Version':'1.0'
}
def get_google_code(secret):
    key = base64.b32decode(secret, True)
    msg = struct.pack(">Q", int(time.time()) // 30)
    google_code = hmac.new(key, msg, hashlib.sha1).digest()
    # 很多网上的代码不可用，就在于这儿，没有chr字符串
    o = ord(chr(google_code[19])) & 15
    # google_code = (struct.unpack(">I", google_code[o:o + 4])[0] & 0x7fffffff) % 1000000
    google_code = (struct.unpack(">I", google_code[o:o + 4])[0] & 0x7fffffff) % 1000000
    return '%06d' % google_code


## python接口提取今日数据
print('启动百度统计API----')
shuju_website = {'domain':[],
                 '日期':[],
                 'pv':[],
                 'uv':[],
                 'ip':[]}
qishi = {'domain':[],
         '日期':[],
         '时间':[],
         'pv':[],
         'uv':[],
         'ip':[]}

# url_siteid = 'https://openapi.baidu.com/rest/2.0/tongji/config/getSiteList?access_token=121.1e832791a57b87542b2bb51e2f3f5bfa.Y_Uhf0W55kh6mBiTGZX0qWg0O5ZqJYZmPyHTqi8.HEyD3w'
# response = requests.get(url_siteid)
#
# dic_website = {}
# for k,v in zip(jsonpath.jsonpath(json.loads(response.text),'$..domain'),jsonpath.jsonpath(json.loads(response.text),'$..site_id')):
#     dic_website[k]=v
with open(r'C:\Users\User\Desktop\SEO\12-18\dic_website.txt','r') as f:
    dic_website=eval(f.read())
# 分别获取各网站数据
app = xw.App(visible=False,add_book=False)
book = app.books.open(r'C:\Users\User\Desktop\SEO\截图文件\今日数据(python接口).xlsx')
sheet1 = book.sheets['网站概况']
# sheet1.range('A2').clear_contents()
sheet_qishu = book.sheets['趋势分析']
# sheet_qishu.range('A2').clear_contents()

session_web = requests.Session()
for k in dic_website:
    url_web = f'https://openapi.baidu.com/rest/2.0/tongji/report/getData?access_token={access_token}&site_id={dic_website[k]}&method=overview/getTimeTrendRpt&start_date={start_date}&end_date={end_date}&metrics=pv_count,visitor_count,ip_count'
    response = session_web.get(url_web)
    response.encoding='utf8'
    # 趋势数据
    for i in range(24):
        qishi['domain'].append(k)
        qishi['日期'].append((datetime.datetime.now()+datetime.timedelta(days=day)).strftime('%Y/%m/%d'))
        qishi['时间'].append(i)
        qishi['pv'].append(json.loads(response.text)['result']['items'][1][i][0])
        qishi['uv'].append(json.loads(response.text)['result']['items'][1][i][1])
        qishi['ip'].append(json.loads(response.text)['result']['items'][1][i][2])
    result_pv_uv_ip = []
    # 遍历列表并相加元素
    for i in range(3):
        sum = 0
        for j in range(len(json.loads(response.text)['result']['items'][1])):
            try:
                sum += json.loads(response.text)['result']['items'][1][j][i]
            except:
                sum +=0
        result_pv_uv_ip.append(sum)

    shuju_website['domain'].append(k)
    shuju_website['日期'].append((datetime.datetime.now()+datetime.timedelta(days=day)).strftime('%Y/%m/%d'))
    shuju_website['pv'].append(result_pv_uv_ip[0])
    shuju_website['uv'].append(result_pv_uv_ip[1])
    shuju_website['ip'].append(result_pv_uv_ip[2])
    time.sleep(1)
sheet1.range('A2').options(index=False,header = False).value = pd.DataFrame(shuju_website)
sheet_qishu.range('A2').options(index=False,header = False).value = pd.DataFrame(qishi)
book.save()
app.quit()
print('今日数据获取完毕！')


# 后续采集会员列表，首充记录、数据处理
# 读取运行jar包的数据，及历史数据
print('读取今日数据。。')
data_today = pd.read_excel(r'C:\Users\User\Desktop\SEO\截图文件\今日数据(python接口).xlsx',index_col=0)
data_2_today = pd.read_excel(r'C:\Users\User\Desktop\SEO\截图文件\今日数据(python接口).xlsx','趋势分析')
daili = pd.read_excel(r'C:\Users\User\Desktop\SEO\数据+ip历史14.xlsx','代理总表')
his_data  = pd.read_excel(r'C:\Users\User\Desktop\SEO\数据+ip历史14.xlsx','数据')

# 采集会员列表和会员存记录
url_fircharge = 'http://fundmng.bsportsadmin.com/api/manage/data/detail/firstRecharge'
url_user = 'http://fundmng.bsportsadmin.com/api/manage/user/maintain/user/list'
session = requests.session()

# 采集首存报表
dic_fir = dict({'会员名':[], '所属代理':[],'注册时间':[], '交易时间':[], '交易类型':[], '币种':[], '金额':[]})
# 昨天开始时间戳
yesterday = datetime.date.today() + datetime.timedelta(days=day)
yesterday_start_time = int(time.mktime(time.strptime(str(yesterday), '%Y-%m-%d')))
# 昨天结束时间戳
yesterday_end_time = int(time.mktime(time.strptime(str(datetime.date.today()), '%Y-%m-%d'))) - 1

#---------------------token----------------------
# 采取token
google_code = get_google_code('64ehnxj6yily5bhv23kgb62ozuh6yuu2')
data0 = {
    'username': 'Marquis',
    'password': 'qwer123456',
    'code': google_code
}
session0 = requests.Session()
response0  =session0.post(url=submit_url,data=data0,headers=header0)
response0.encoding = 'utf-8'
obj0 = json.loads(response0.text)
token = obj0['data']['token']

header = {
    'Accept':'application/json, text/plain, */*',
    # 'Accept-Encoding':'gzip, deflate',
    'Accept-Language':'zh-CN,zh;q=0.9',
    'Connection':'keep-alive',
    'Content-Length':'75',
    'Content-Type':'application/x-www-form-urlencoded',
    'Cookie':'admin-token=67c8b1bd1b434f898ed8570a860355b8; admin-uid=690',
    'Device_id':'1.0',
    'Gl_version':'2.0',
    'Host':'fundmng.bsportsadmin.com',
    'Language':'zh_CN',
    'Menuid':'100112',
    'Opeartionmenu':'%u62A5%u8868%u67E5%u8BE2-%u4F1A%u5458%u9996%u5B58%u62A5%u8868',
    'Origin':'http://fundmng.bsportsadmin.com',
    'Os_type':'0',
    'Referer':'http://fundmng.bsportsadmin.com/system/report-query/report-first-recharge',
    'Sign':'ca83944852acc68fe114cbc65f1e1d22',
    'Some':'header',
    'Systemid':'54',
    'Timestamp':'1692092554000',
    'Token':token,
    'Uid':'690',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36',
    'Version':'1.0'
}
for page in range(1,pages_fircharge+1):
    data = {
        'page': page,
        'size': 20,
        'tradeType': 0,
        'isFake': 0,
        'stime': yesterday_start_time*1000,
        'etime': yesterday_end_time*1000+999
    }
    response = session.post(url_fircharge,headers=header,data=data)
    response.encoding = 'utf-8'
    obj = json.loads(response.text)

    for i in obj['data']['list']:
        dic_fir['会员名'].append(i['userName'])
        dic_fir['所属代理'].append(i['parentName'])
        dic_fir['注册时间'].append(time.strftime('%Y/%m/%d %H:%M:%S',time.localtime(i['createTime']//1000)))
        dic_fir['交易时间'].append(time.strftime('%Y/%m/%d %H:%M:%S',time.localtime(i['regTime']//1000)))
        dic_fir['交易类型'].append(i['tradeType'])
        dic_fir['币种'].append(i['coinCode'])
        dic_fir['金额'].append(i['amount'])
firChargeUser = pd.DataFrame(dic_fir)
print('会员首存行列：',firChargeUser.shape)

# 采集会员列表
dic_user = dict({'会员账号':[], '姓名':[],'代理':[], '注册时间':[], '备注':[]})
for page in range(1,pages_user+1):
    data2 = {
        'page':page,
        'size':20,
        'userVip':'0,1,2,3,4,5,6,7,8,9,10,11',
        'status':'0,1,2,4',
        'sortType':'3',
        'sortStr':'descend',
        'searchType':'1',
        'channelId':'34',
        'registeredStartDate':yesterday_start_time*1000,
        'registeredEndDate':yesterday_end_time*1000+999,
    }
    response2 = session.post(url_user,headers=header,data=data2)
    response2.encoding = 'utf-8'
    obj2 = json.loads(response2.text)
    for i in obj2['data']['list']:
        dic_user['会员账号'].append(i['username'])
        if i['reallyName'] !='':
            dic_user['姓名'].append(i['reallyName'])
        else:
            dic_user['姓名'].append('--')

        dic_user['代理'].append(i['parentName'])
        dic_user['注册时间'].append(time.strftime('%Y/%m/%d %H:%M:%S',time.localtime(i['registerDate']//1000)))
        dic_user['备注'].append(i['remark'])
user = pd.DataFrame(dic_user)
print('用户列表行列:',user.shape)
# 删除测试账号
user = user[~user['会员账号'].str.contains('test')&~user['会员账号'].str.contains('ceshi')&~user['姓名'].str.contains('测试') \
            &~user['姓名'].str.contains('cheshi')&~user['代理'].str.contains('测试')&~user['代理'].str.contains('cheshi') \
            &~user['备注'].str.contains('测试')&~user['备注'].str.contains('试玩')&~user['备注'].str.contains('晒单')]
print('去重后：',user.shape)

#  开始处理数据---------------------------------------------
print('开始处理shuju')
shuju = pd.DataFrame({'人员':['Paddy', 'Tony', 'Max', 'Martin', 'Zed', 'Hugo', 'Aber', 'DK', 'Ben','当日汇总'],
                      '日期':(datetime.datetime.now()+datetime.timedelta(days=day)).strftime('%Y/%m/%d'),
                      '发送IP':0,
                      '接受IP':0,
                      '对比昨天(总IP)':0,
                      '对比前3天均值(总IP)':0,
                      '对比前5天均值(总IP)':0,
                      '对比前7天均值(总IP)':0,
                      '对比昨天(总注册)':0,
                      '对比前3天均值(总注册)':0,
                      '对比前5天均值(总注册)':0,
                      '对比前7天均值(总注册)':0,
                      '对比昨天(总开户)':0,
                      '对比前3天均值(总开户)':0,
                      '对比前5天均值(总开户)':0,
                      '对比前7天均值(总开户)':0})


shuju.set_index('人员',inplace = True)

# data_today['IP']=pd.to_numeric(data_today['IP'],errors='coerce').replace(np.nan,0).astype('int64')
# data_today=data_today.groupby('网站名(domain)').agg({'IP':sum})
try:
    shuju.loc['Paddy','发送IP']=data_today.loc['paddy.com','IP']
except:
    shuju.loc['Paddy','发送IP']=0
try:
    shuju.loc['Paddy','接受IP']=data_today.loc['paddy.bty','IP']
except:
    shuju.loc['Paddy','接受IP']=0
try:
    shuju.loc['Tony','发送IP']=data_today.loc['tonyb.com','IP']
except:
    shuju.loc['Tony', '发送IP'] = 0
try:
    shuju.loc['Tony','接受IP']=data_today.loc['tony.bty','IP']
except:
    shuju.loc['Tony', '接受IP'] = 0
try:
    shuju.loc['Max','发送IP']=data_today.loc['mulu.com','IP']
except:
    shuju.loc['Max', '发送IP'] = 0
try:
    shuju.loc['Max','接受IP']=data_today.loc['max.bty','IP']
except:
    shuju.loc['Max', '接受IP'] = 0
try:
    shuju.loc['Martin','发送IP']=data_today.loc['redquan.com','IP']
except:
    shuju.loc['Martin', '发送IP'] = 0
try:
    shuju.loc['Martin','接受IP']=data_today.loc['martin.bty','IP']
except:
    shuju.loc['Martin','接受IP']=0
try:
    shuju.loc['Zed','发送IP']=data_today.loc['zed.com','IP']
except:
    shuju.loc['Zed', '发送IP'] = 0
try:
    shuju.loc['Zed','接受IP']=data_today.loc['zed.bty','IP']
except:
    shuju.loc['Zed', '接受IP'] = 0
try:
    shuju.loc['Hugo','发送IP']=data_today.loc['hugo.com','IP']
except:
    shuju.loc['Hugo', '发送IP'] = 0
try:
    shuju.loc['Hugo','接受IP']=data_today.loc['hugo.bty','IP']
except:
    shuju.loc['Hugo','接受IP']=0
try:
    shuju.loc['Aber','发送IP']=data_today.loc['aber.com','IP']
except:
    shuju.loc['Aber','发送IP']=0
try:
    shuju.loc['Aber','接受IP']=data_today.loc['aber.bty','IP']
except:
    shuju.loc['Aber','接受IP']=0
shuju.loc['DK','发送IP']=data_today.loc['dk.com','IP']
shuju.loc['DK','接受IP']=data_today.loc['dk.bty','IP']
shuju.loc['Ben','发送IP']=data_today.loc['ben.com','IP']
shuju.loc['Ben','接受IP']=data_today.loc['ben.bty','IP']
shuju.loc['当日汇总','发送IP']=shuju['发送IP'].sum()
shuju.loc['当日汇总','接受IP']=shuju['接受IP'].sum()

shuju['日期'] = pd.to_datetime(shuju['日期'])
shuju.insert(1,'人员2',shuju.index)
shuju['人员2']=shuju['人员2'].str.lower()

# 第1次merge前，重置索引
shuju.reset_index(inplace=True)

merge_user = pd.merge(user,daili,how = 'left',left_on='代理',right_on='代理线')
data_todaySEO = merge_user.groupby('seo变化数据团队').agg({'seo变化数据团队':len})
data_todaySEO.rename(columns={'seo变化数据团队':'注册'},inplace=True)
data_todaySEO.reset_index(inplace=True)
data_todaySEO['人员2'] = data_todaySEO['seo变化数据团队'].str.lower()
data_todaySEO.set_index('seo变化数据团队',inplace=True)

shuju=shuju.merge(data_todaySEO,on='人员2',how='left')

shuju['注册率(%)'] = round(shuju['注册']/shuju['发送IP']*100,2)

merge_charge = pd.merge(firChargeUser,daili,how='left',left_on='所属代理',right_on='代理线')
data_todayCHARGE = merge_charge.groupby('seo变化数据团队').agg({'seo变化数据团队':len})
data_todayCHARGE= data_todayCHARGE.rename(columns={'seo变化数据团队':'开户'})
data_todayCHARGE.reset_index(inplace=True)
data_todayCHARGE['seo变化数据团队']=data_todayCHARGE['seo变化数据团队'].str.lower()
data_todayCHARGE= data_todayCHARGE.rename(columns={'seo变化数据团队':'人员2'})
# 第2次merge
shuju = pd.merge(shuju,data_todayCHARGE,how='left',on='人员2')
shuju['转化率(%)'] = round(shuju['开户']/shuju['注册']*100,2)

merge_charge['注册时间']= pd.to_datetime(merge_charge['注册时间'])
merge_charge['交易时间']= pd.to_datetime(merge_charge['交易时间'])
data_today3  = merge_charge[merge_charge['注册时间'].dt.strftime('%Y/%m/%d')==merge_charge['交易时间'].dt.strftime('%Y/%m/%d')].groupby('seo变化数据团队').agg({'seo变化数据团队':len})
data_today3.rename(columns = {'seo变化数据团队':'当日注册并开户'},inplace=True)
data_today3.reset_index(inplace=True)
data_today3['seo变化数据团队'] =data_today3['seo变化数据团队'].str.lower()
data_today3.rename(columns = {'seo变化数据团队':'人员2'},inplace=True)
# 第3次merge
shuju  = pd.merge(shuju,data_today3,how='left',on='人员2')
shuju['当日注册激活率(%)'] = round(shuju['当日注册并开户']/shuju['注册']*100,2)


#------------
# his_data  = pd.read_csv(r'C:\Users\User\Desktop\SEO\SEO每日更新_814.csv',encoding='gbk')
his_data['日期']= pd.to_datetime(his_data['日期'])
be_data = his_data[his_data['日期']==(shuju['日期'][0]+datetime.timedelta(days=-1))][:-1]

shuju.fillna(0,inplace=True)
shuju.set_index('人员',inplace = True)
shuju.sort_index(inplace=True)
be_data.set_index('人员',inplace=True)
be_data.sort_index(inplace=True)

be3_data = his_data[his_data['日期']>=(shuju['日期'][0]+datetime.timedelta(days=-3))]
be3_data = be3_data.groupby('人员').mean()[:-1]
be5_data = his_data[his_data['日期']>=(shuju['日期'][0]+datetime.timedelta(days=-5))]
be5_data = be5_data.groupby('人员').mean()[:-1]
be7_data = his_data[his_data['日期']>=(shuju['日期'][0]+datetime.timedelta(days=-7))]
be7_data = be7_data.groupby('人员').mean()[:-1]

shuju['对比昨天(总IP)']=shuju['发送IP']-be_data['总IP']

shuju['对比前3天均值(总IP)']= shuju['发送IP']-be3_data['总IP']
shuju['对比前5天均值(总IP)']= shuju['发送IP']-be5_data['总IP']
shuju['对比前7天均值(总IP)']= shuju['发送IP']-be7_data['总IP']

shuju['对比昨天(总注册)']=shuju['注册']-be_data['注册']
shuju['对比前3天均值(总注册)']= shuju['注册']-be3_data['注册']
shuju['对比前5天均值(总注册)']= shuju['注册']-be5_data['注册']
shuju['对比前7天均值(总注册)']= shuju['注册']-be7_data['注册']

shuju['对比昨天(总开户)']=shuju['开户']-be_data['开户']
shuju['对比前3天均值(总开户)']= shuju['开户']-be3_data['开户']
shuju['对比前5天均值(总开户)']= shuju['开户']-be5_data['开户']
shuju['对比前7天均值(总开户)']= shuju['开户']-be7_data['开户']

shuju = shuju.iloc[:,:4].join(shuju.iloc[:,-6:]).join(shuju.iloc[:,4:-6])
shuju.fillna(0,inplace=True)
for name in shuju.index:
    if shuju.loc[name,'注册']==0:
        shuju.loc[name,'转化率(%)']=shuju.loc[name,'开户']*100
        shuju.loc[name,'当日注册激活率(%)']=shuju.loc[name,'当日注册并开户']*100
    if shuju.loc[name,'发送IP']==0:
        shuju.loc[name,'注册率(%)']=shuju.loc[name,'接受IP']*100

shuju.loc[:,'对比昨天(总IP)':'对比前7天均值(总开户)']=shuju.loc[:,'对比昨天(总IP)':'对比前7天均值(总开户)'].astype('int64')
shuju['注册'] = shuju['注册'].astype('int64')
shuju['开户'] = shuju['开户'].astype('int64')
shuju['当日注册并开户'] = shuju['当日注册并开户'].astype('int64')


for i in shuju.iloc[:,4:].columns:
    shuju.loc['当日汇总',i]=shuju[i].sum()
# 重置三个率
shuju.loc['当日汇总','注册率(%)']=round(shuju.loc['当日汇总','注册']/shuju.loc['当日汇总','发送IP']*100,2)
shuju.loc['当日汇总','转化率(%)']=round(shuju.loc['当日汇总','开户']/shuju.loc['当日汇总','注册']*100,2)
shuju.loc['当日汇总','当日注册激活率(%)']=round(shuju.loc['当日汇总','当日注册并开户']/shuju.loc['当日汇总','注册']*100,2)

shuju.insert(1,'人员',shuju.index)
shuju.drop('人员2',inplace=True,axis=1)
print('shuju处理完成。。。。')

# 开始计算ip历史数据
print('开始计算ip历史数据。。。。')

dic_ip ={'日期':(datetime.datetime.now()+datetime.timedelta(days=day)).strftime('%Y/%m/%d'),
         '人员':[i for i in ['Paddy', 'Tony', 'Max', 'Martin', 'Zed', 'Hugo', 'Aber', 'DK', 'Ben'] for j in range(7)],'指标':['发送IP数','接收IP数','接收率(%)','注册','注册率(%)','开户','开户转化率(%)']*9, '总计':0, '0-2':0, '2-4':0, '4-6':0, '6-8':0, '8-10':0, '10-12':0, '12-14':0, '14-16':0, '16-18':0, '18-20':0, '20-22':0, '22-24':0}
ip_data = pd.DataFrame(dic_ip)
user['注册时间']=pd.to_datetime(user['注册时间'])
hour_user= pd.merge(user,daili,how = 'left',left_on='代理',right_on='代理线')
hour_user['小时数']= hour_user['注册时间'].dt.hour
hour_user['seo变化数据团队'] = hour_user['seo变化数据团队'].str.lower()
firChargeUser['注册时间']=pd.to_datetime(firChargeUser['注册时间'])
hour_charge = pd.merge(firChargeUser,daili,how='left',left_on='所属代理',right_on='代理线')
hour_charge['小时数']= hour_charge['注册时间'].dt.hour
hour_charge['seo变化数据团队'] = hour_charge['seo变化数据团队'].str.lower()

data_2_today['PV'] = pd.to_numeric(data_2_today['PV'],errors='coerce',downcast='integer')
data_2_today['UV'] = pd.to_numeric(data_2_today['UV'],errors='coerce',downcast='integer')
data_2_today['IP'] = pd.to_numeric(data_2_today['IP'],errors='coerce',downcast='integer')

# 循环方式
name_list = ['Martin','Paddy', 'Tony', 'Max',  'Zed', 'Hugo', 'Aber', 'DK', 'Ben']
hour_list = ['0-2', '2-4', '4-6', '6-8', '8-10', '10-12', '12-14', '14-16', '16-18', '18-20', '20-22', '22-24']
web_dic={'Martin':['redquan.com','martin.bty'],
         'Paddy':['paddy.com','paddy.bty'],
         'Tony':['tonyb.com','tony.bty'],
         'Max':['mulu.com','max.bty'],
         'Zed':['zed.com','zed.bty'],
         'Hugo':['hugo.com','hugo.bty'],
         'Aber':['aber.com','aber.bty'],
         'DK':['dk.com','dk.bty'],
         'Ben':['ben.com','ben.bty']}

for name in name_list:
    for h in hour_list:
        ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='发送IP数'),h] =data_2_today[data_2_today['网站名(domain)'].str.contains(web_dic[name][0])&(data_2_today['时间']>=int(h.split('-')[0])) & (data_2_today['时间']<int(h.split('-')[1]))]['IP'].sum()
        ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='接收IP数'),h] =data_2_today[data_2_today['网站名(domain)'].str.contains(web_dic[name][1])&(data_2_today['时间']>=int(h.split('-')[0])) & (data_2_today['时间']<int(h.split('-')[1]))]['IP'].sum()
        ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='注册'),h] =len(hour_user[(hour_user['seo变化数据团队']==name.lower()) & (hour_user['小时数']>=int(h.split('-')[0]))& (hour_user['小时数']<int(h.split('-')[1]))])
        ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='注册率(%)'),h]=round(len(hour_user[(hour_user['seo变化数据团队']==name.lower()) & (hour_user['小时数']>=int(h.split('-')[0]))& (hour_user['小时数']<int(h.split('-')[1]))])/ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='接收IP数'),h].iloc[0]*100,2)
        ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='开户'),h] =len(hour_charge[(hour_charge['seo变化数据团队']==name.lower()) & (hour_charge['小时数']>=int(h.split('-')[0]))& (hour_charge['小时数']<int(h.split('-')[1]))])
        ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='开户转化率(%)'),h]=round(len(hour_charge[(hour_charge['seo变化数据团队']==name.lower()) & (hour_charge['小时数']>=int(h.split('-')[0]))& (hour_charge['小时数']<int(h.split('-')[1]))])/ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='注册'),h].iloc[0]*100,2)
        ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='接收率(%)'),h] =round(ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='接收IP数'),h].iloc[0] / ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='发送IP数'),h].iloc[0]*100,2)

    ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='发送IP数'),'总计'] =ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='发送IP数'),'0-2':].T.sum()
    ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='接收IP数'),'总计'] =ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='接收IP数'),'0-2':].T.sum()
    ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='注册'),'总计'] =ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='注册'),'0-2':].T.sum()
    ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='开户'),'总计'] =ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='开户'),'0-2':].T.sum()
    ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='开户转化率(%)'),'总计'] =round(ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='开户'),'总计'].iloc[0] / ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='注册'),'总计'].iloc[0]*100,2)
    ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='注册率(%)'),'总计'] =round(ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='注册'),'总计'].iloc[0] / ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='接收IP数'),'总计'].iloc[0]*100,2)
    ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='接收率(%)'),'总计'] =round(ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='接收IP数'),'总计'].iloc[0] / ip_data.loc[(ip_data['人员']==name) & (ip_data['指标']=='发送IP数'),'总计'].iloc[0]*100,2)
# 增加行末表头
header_shuju = pd.DataFrame({'人员':'人员',
                             '日期':'日期',
                             '发送IP':'总IP',
                             '接受IP':'接受IP',
                             '注册':'注册',
                             '注册率(%)':'注册率(%)',
                             '开户':'开户',
                             '转化率(%)':'转化率(%)',
                             '当日注册并开户':'当日注册并开户',
                             '当日注册激活率(%)':'当日注册激活率(%)',
                             '对比昨天(总IP)':'对比昨天(总IP)',
                             '对比前3天均值(总IP)':'对比前3天均值(总IP)',
                             '对比前5天均值(总IP)':'对比前5天均值(总IP)',
                             '对比前7天均值(总IP)':'对比前7天均值(总IP)',
                             '对比昨天(总注册)':'对比昨天(总注册)',
                             '对比前3天均值(总注册)':'对比前3天均值(总注册)',
                             '对比前5天均值(总注册)':'对比前5天均值(总注册)',
                             '对比前7天均值(总注册)':'对比前7天均值(总注册)',
                             '对比昨天(总开户)':'对比昨天(总开户)',
                             '对比前3天均值(总开户)':'对比前3天均值(总开户)',
                             '对比前5天均值(总开户)':'对比前5天均值(总开户)',
                             '对比前7天均值(总开户)':'对比前7天均值(总开户)'},index=[0])
with open(r'C:\Users\User\Desktop\SEO\截图文件\seo_全天.txt','w') as f:
    f.write(f'#SEO数据  {(datetime.datetime.now()+datetime.timedelta(days=day)).strftime("%Y/%m/%d")}\n')
    f.write(f'转化率<30%的人员：{str(list(shuju[:-1].loc[shuju["转化率(%)"]<30,:]["人员"]))}\n')
    f.write(f'较前一天总IP下降人员为：{str(list(shuju[:-1].loc[shuju["对比昨天(总IP)"]<0,:]["人员"]))}')
# 增加%
shuju['注册率(%)'] =shuju['注册率(%)'].apply(lambda x: str(x)+'%')
shuju['转化率(%)'] =shuju['转化率(%)'].apply(lambda x: str(x)+'%')
shuju['当日注册激活率(%)'] =shuju['当日注册激活率(%)'].apply(lambda x: str(x)+'%')
shuju = shuju.append(header_shuju)
header_ip =pd.DataFrame({'日期':'日期',
                         '人员':'人员','指标':'指标', '总计':'总计', '0-2':'0-2时', '2-4':'2-4时', '4-6':'4-6时', '6-8':'6-8时', '8-10':'8-10时', '10-12':'10-12时', '12-14':'12-14时', '14-16':'14-16时', '16-18':'16-18时', '18-20':'18-20时', '20-22':'20-22时', '22-24':'22-24时'},index=[0])
ip_data= ip_data.append(header_ip)
print(shuju)
#----------------------------------------------
ip_DATA= pd.DataFrame()
for name in set(ip_data.iloc[:-1,:]['人员']):
    ip_DATA = ip_DATA.append(ip_data.loc[ip_data['人员']==name,:].append(ip_data.iloc[-1,:]))

# 更新每日数据--------------------------------------------------------------------------------------------------
app = xw.App(visible=False,add_book=False)
book = app.books.open(r'C:\Users\User\Desktop\SEO\数据+ip历史14.xlsx')

sheet_shuju = book.sheets['数据']
row_shuju = sheet_shuju.used_range.last_cell.row

sheet_ip =  book.sheets['ip历史']
row_ip = sheet_ip.used_range.last_cell.row

sheet_shuju['A'+str(row_shuju+1)].options(index=False,header = False).value = shuju
sheet_ip['A'+str(row_ip+1)].options(index=False,header = False).value = ip_DATA
book.save()
book.close()
#
# # 添加条件格式
wb = load_workbook(r'C:\Users\User\Desktop\SEO\数据+ip历史14.xlsx')
ws = wb['数据']
# redFill = PatternFill(start_color='EE1111',end_color='EE1111',fill_type='solid')
redFill = Font(color='FF0000')
# ws.conditional_formatting.add(f'K{row_shuju-9}:V{row_shuju}',
#                 formatting.rule.CellIsRule(operator='lessThan',
#                 formula=['0'],
#                 stopIfTrue=True,
#                 fill=redFill))
ws.conditional_formatting.add(f'K{row_shuju +1}:V{row_shuju +10}',
                              formatting.rule.CellIsRule(operator='lessThan',
                                                         formula=['0'],
                                                         font=redFill))
# ip历史增加颜色
ws_ip = wb['ip历史']
source_range = ws_ip[f'A{row_ip-72}:P{row_ip-1}']
# 复制源区域的单元格格式到目标区域
for row in source_range:
    for cell in row:
        # 获取目标单元格
        target_cell = ws_ip.cell(row=cell.row+72, column=cell.column)
        # 复制单元格格式
        target_cell.font = cell.font.copy()
        target_cell.border = cell.border.copy()
        target_cell.fill = cell.fill.copy()
        target_cell.number_format = cell.number_format
        target_cell.protection = cell.protection.copy()
        target_cell.alignment = cell.alignment.copy()
# 保存工作簿
wb.save(filename=r'C:\Users\User\Desktop\SEO\数据+ip历史14.xlsx')
wb.close()
# 保存截图
# pyperclip.copy('')
book2 = app.books.open(r'C:\Users\User\Desktop\SEO\数据+ip历史14.xlsx')
s_book = app.books.open(rf'C:\Users\User\Desktop\SEO\SEO输出(9点)\SEO数据_{last_date}.xlsx')
s_sheet1 = s_book.sheets['Sheet1']
s_sheet2 = s_book.sheets['Sheet2']
sheet2_shuju = book2.sheets['数据']
sheet2_ip =  book2.sheets['ip历史']
sheet_tem = book2.sheets['临时']
# 复制源Excel的区域到目标Excel的区域
source_range = sheet2_shuju.range(f'A{row_shuju+1}:V{row_shuju+10}')

target_range = sheet_tem.range('A3:V12')
source_range.copy()
target_range.paste()
book2.save()
# 粘贴至发送表格
s_sheet1.range('A2:v11').paste()

# 复制图片
# pyperclip.copy('')
range_shuju = sheet_tem.range('A1:V12')
range_shuju.api.CopyPicture()
img_shuju = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
img_shuju.save(r'C:\Users\User\Desktop\SEO\截图文件\shuju.png')  # 保存图片
# 删除行末表头

def delete_row(sheet, row_index):
    range_obj = sheet.range(f'A{row_index}:A{row_index}')
    range_obj.api.EntireRow.Delete()
delete_row(sheet2_shuju,row_shuju+11)
time.sleep(2)

range_IP = sheet2_ip.range(f'A{row_ip}:P{row_ip+71}')
#粘贴到发送表格
range_IP.copy()
s_sheet2.range('A1:P72').paste()
s_book.save(fr'C:\Users\User\Desktop\SEO\SEO输出(9点)\SEO数据_{start_date}.xlsx')
range_IP.api.CopyPicture()
img_IP = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
img_IP.save(r'C:\Users\User\Desktop\SEO\截图文件\IP.png')  # 保存图片

time.sleep(2)
book2.save()
book2.close()
app.quit()
#新增加发送表格

# 发送到群
with open(r'C:\Users\User\Desktop\SEO\截图文件\seo_全天.txt','r') as f:
    text = f.read()
bot_DA = telebot.TeleBot("6106076754:AAHjxPSBpyjwpY-lq1iEslUufW46XQvAfr0")
# bot_m = telebot.TeleBot("6377312623:AAGz3ZSMVswWq0QVlihRPklw8b7skSBP16Y") -812533282  -677235937  "鲲鹏": -321785338
bot_DA.send_photo(-677235937,open(r'C:\Users\User\Desktop\SEO\截图文件\shuju.png','rb'),timeout=100)
bot_DA.send_message(-677235937,text,timeout=100)
bot_DA.send_photo(-677235937,open(r'C:\Users\User\Desktop\SEO\截图文件\IP.png','rb'),timeout=100)
# bot_DA.send_document(-677235937,open(r"C:\Users\User\Desktop\SEO\数据+ip历史14.xlsx",'rb'),timeout=600)
bot_DA.send_document(-677235937,open(rf'C:\Users\User\Desktop\SEO\SEO输出(9点)\SEO数据_{start_date}.xlsx','rb'),timeout=600)
bot_DA.stop_polling()
# 查看
# print(shuju)
# print(ip_data)